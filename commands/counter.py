import command_system
import vk
from settings import s_token, token
from openpyxl import load_workbook
from flask_app import curr_user_id
import time
import random

session = vk.Session()
api = vk.API(session, v=5.92)

def counter():
    #posts = api.wall.search(owner_id=-177456743, domain = 'test_bot_df', query = '#событие', count = 1, access_token = s_token)
    #message = posts['items'][0]['text']
    user_id = curr_user_id()
    hist = api.messages.getHistory(peer_id = user_id, group_id=21167006, offset=0, count =1, access_token = token)
    cand = hist['items'][0]['text']
    protect = api.messages.getHistory(peer_id = user_id, group_id=21167006, offset=2, count =1, access_token = token)['items'][0]['text']
    if protect=='Наташа Ж.' or protect=='Карина Ч.' or protect=='Марина П.' or protect=='Настя К.' or protect=='Валя Ч.' or protect=='Катя К.' or protect=='Ульяна М.' or protect=='Виолетта П.' or protect=='Лера Л.' or protect=='Наташа У.':
        message='Ты уже голосовал!'
        buttons=None
        return message, '', buttons
    message='Спасибо! Твой голос учтен.'
    time.sleep(random.randint(0,10))
    wb = load_workbook('/home/votebot/mysite/commands/voting.xlsx')
    sheet = wb.active
    b2=sheet.cell(row=2, column=2).value
    b3=sheet.cell(row=3, column=2).value
    b4=sheet.cell(row=4, column=2).value
    b5=sheet.cell(row=5, column=2).value
    b6=sheet.cell(row=6, column=2).value
    b7=sheet.cell(row=7, column=2).value
    b8=sheet.cell(row=8, column=2).value
    b9=sheet.cell(row=9, column=2).value
    b10=sheet.cell(row=10, column=2).value
    b11=sheet.cell(row=11, column=2).value
    if cand=='Наташа Ж.':
        b2+=1
        sheet.cell(row=2, column=2).value=b2
    elif cand=='Карина Ч.':
        b3+=1
        sheet.cell(row=3, column=2).value=b3
    elif cand=='Марина П.':
        b4+=1
        sheet.cell(row=4, column=2).value=b4
    elif cand=='Настя К.':
        b5+=1
        sheet.cell(row=5, column=2).value=b5
    elif cand=='Валя Ч.':
        b6+=1
        sheet.cell(row=6, column=2).value=b6
    elif cand=='Катя К.':
        b7+=1
        sheet.cell(row=7, column=2).value=b7
    elif cand=='Ульяна М.':
        b8+=1
        sheet.cell(row=8, column=2).value=b8
    elif cand=='Виолетта П.':
        b9+=1
        sheet.cell(row=9, column=2).value=b9
    elif cand=='Лера Л.':
        b10+=1
        sheet.cell(row=10, column=2).value=b10
    elif cand=='Наташа У.':
        b11+=1
        sheet.cell(row=11, column=2).value=b11
    wb.save('/home/votebot/mysite/commands/voting.xlsx')
    #message = str(b2)+' '+str(b3)+' '+str(b4)
    buttons = None
    return message, '', buttons

counter_command = command_system.Command()

counter_command.keys = ['Наташа Ж.', 'Карина Ч.','Марина П.','Настя К.','Валя Ч.','Катя К.','Ульяна М.','Виолетта П.','Лера Л.','Наташа У.']
counter_command.description = 'кандидаты'
counter_command.process = counter

